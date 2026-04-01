#!/usr/bin/env python3
"""
HSI v11 - Excel Workbook Generator
Combines all 7 CSV tracking files into a single Excel workbook with multiple sheets
"""

import csv
import os
from datetime import datetime

def csv_to_excel_sheet(csv_file, sheet_name):
    """Read CSV and return data for Excel sheet"""
    data = []
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            data.append(row)
    return data

def create_excel_workbook():
    """Create XLSX-style workbook using openpyxl if available, otherwise create formatted CSV bundle"""
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.utils.dataframe import dataframe_to_rows
        use_openpyxl = True
    except ImportError:
        print("⚠️ openpyxl not available, creating enhanced CSV bundle instead")
        use_openpyxl = False
    
    # List of CSV files to combine
    csv_files = [
        ('hsi_v11_dashboard.csv', 'Dashboard'),
        ('hsi_v11_cycle_tracker.csv', 'Cycle Tracker'),
        ('hsi_v11_sector_allocation.csv', 'Sector Allocation'),
        ('hsi_v11_solar_calendar_2026.csv', 'Solar Calendar 2026'),
        ('hsi_v11_convergence_history.csv', 'Convergence History'),
        ('hsi_v11_v10_integration.csv', 'V10 Integration'),
        ('hsi_v11_constituents_watchlist.csv', 'Constituents Watchlist'),
    ]
    
    if use_openpyxl:
        # Create Excel workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Define styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='006699', end_color='006699', fill_type='solid')
        title_font = Font(bold=True, size=14, color='006699')
        cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for csv_file, sheet_name in csv_files:
            if not os.path.exists(csv_file):
                print(f"⚠️ {csv_file} not found, skipping...")
                continue
            
            ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit: 31 chars
            
            # Read CSV data
            data = csv_to_excel_sheet(csv_file, sheet_name)
            
            # Write data to sheet
            for row_idx, row_data in enumerate(data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Apply header styling to first row
                    if row_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = cell_alignment
                        cell.border = thin_border
                    else:
                        cell.border = thin_border
                        cell.alignment = cell_alignment
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
        
        # Save workbook
        output_file = 'HSI_v11_Complete_Tracker.xlsx'
        wb.save(output_file)
        print(f"✅ Excel workbook created: {output_file}")
        print(f"   Sheets: {len(wb.sheetnames)}")
        for sheet in wb.sheetnames:
            print(f"   - {sheet}")
    
    else:
        # Create enhanced CSV bundle with formatting instructions
        output_file = 'HSI_v11_Complete_Tracker.csv'
        with open(output_file, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            
            for csv_file, sheet_name in csv_files:
                if not os.path.exists(csv_file):
                    continue
                
                # Add sheet separator
                writer.writerow(['=' * 80])
                writer.writerow([f'SHEET: {sheet_name}'])
                writer.writerow(['=' * 80])
                writer.writerow([])
                
                # Copy CSV content
                with open(csv_file, 'r', encoding='utf-8') as src:
                    reader = csv.reader(src)
                    for row in reader:
                        writer.writerow(row)
                
                writer.writerow([])
                writer.writerow([])
        
        print(f"✅ CSV bundle created: {output_file}")
        print("   (Import into Excel and use Text-to-Columns to separate sheets)")
    
    print("\n📊 Next steps:")
    print("   1. Open the file in Excel/Google Sheets")
    print("   2. Apply conditional formatting for signal colors")
    print("   3. Set up data validation for status fields")
    print("   4. Save as your master tracking workbook")

if __name__ == '__main__':
    create_excel_workbook()
